import os
import sys
import time
import unittest
import tempfile
from pathlib import Path
from openpyxl import Workbook

# Set project_root by moving one level up from tests and add src to sys.path.
project_root = Path(__file__).resolve().parent.parent
src_dir = project_root / "src"
sys.path.insert(0, str(src_dir))
print("Added to sys.path:", src_dir)

from data_processing.excel_preprocessor import ExcelPreprocessor


class TestDeMerge(unittest.TestCase):
    def setUp(self):
        # Create a temporary Excel file with a merged cell using openpyxl.
        self.temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb = Workbook()
        ws = wb.active
        # Merge cells A1 and B1 and set the value "TestHeader"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws["A1"] = "TestHeader"
        # Add additional data to row 2.
        ws["A2"] = "Data1"
        ws["B2"] = "Data2"
        wb.save(self.temp_file.name)
        wb.close()
        print("Temporary Excel file created at:", self.temp_file.name)

    def tearDown(self):
        # Wait a short moment to allow any file locks to release.
        time.sleep(0.5)
        try:
            os.unlink(self.temp_file.name)
            print("Temporary file removed:", self.temp_file.name)
        except Exception as e:
            print(f"Error cleaning up temporary file: {e}")

    def test_demerge(self):
        preprocessor = ExcelPreprocessor()
        # Process the temporary file.
        processed_data = preprocessor.process_excel_full_range(
            file_path=self.temp_file.name,
            yearbook_source="TestSource",
            group="TestGroup",
        )
        # Verify that data was processed.
        self.assertGreater(len(processed_data), 0, "No data was processed.")
        first_row = processed_data[0]
        # The processed row should have a 'text' field with comma-separated cell values.
        cells = [cell.strip() for cell in first_row["text"].split(",")]
        print("Processed cells:", cells)
        # Assert that "TestHeader" appears at least twice in the merged region.
        # Adjust the expected count if your de-merging behavior changes.
        self.assertGreaterEqual(
            cells.count("TestHeader"),
            2,
            "The merged cell value was not duplicated as expected.",
        )


if __name__ == "__main__":
    unittest.main(exit=False)
